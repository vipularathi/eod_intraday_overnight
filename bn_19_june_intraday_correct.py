import time
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import datetime
import mibian
import os
import glob
import sqlalchemy as sql
import seaborn as sns
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import sys
import yfinance as yf
import plotly.graph_objects as go

from assets import ceiling_xcl, get_greeks, calc_business_days, sumproduct, change_format
from common import root_dir, logs_dir, logger, today_date, yesterday, tomorrow
from remote_db_ops import from_master

ticker, sym_list, exp_list = [], [], []
csv_file = None
df_fut = pd.DataFrame()

def get_ce_pe_price(row_index, opt_type):
    # ce_ticker = row['CE']
    if opt_type == 'CE':
        ticker = df_fut.loc[row_index, 'ce']
    if opt_type == 'PE':
        ticker = df_fut.loc[row_index, 'pe']
    time = df_fut.loc[row_index, 'time']
    price = df.loc[(df['Ticker'] == ticker) & (df['Time'] == time), 'Close'].values[0]
    return price
def save_to_sheet(sym, exp, df, sheet_name, wb, plot_graph = False):
    # if sheet_name == 'intraday':
    ws = wb.create_sheet(title=sheet_name)
    for each_line in dataframe_to_rows(df, index=False, header=True):
        print(each_line)
        ws.append(each_line)
# ----------------------------------------------------------------
    if plot_graph:
        # stt = '09:19:59'
        # ett = '15:19:59'
        # filtered_df = df[(df['time'] >= stt) & (df['time'] <= ett)]
        #
        # sns.set(style='whitegrid')
        # sns.set_context('talk')
        #
        # # fig, ax1 = plt.subplot(figsize=(12,6))
        # fig, ax1 = plt.subplots(figsize=(15, 7.5))
        #
        # sns.lineplot(x='time', y='final_pnl', data=filtered_df, ax=ax1, color='blue', label='Final PnL')
        # ax1.set_ylabel('Final PnL', color='blue', fontsize=8)
        # ax1.tick_params(axis='y', labelcolor='blue')
        #
        # ax2 = ax1.twinx()
        #
        # sns.lineplot(x='time', y='forward', data=filtered_df, ax=ax2, color='orange', label='Forward')
        # ax2.set_ylabel('Forward', color='orange', fontsize=5)
        # ax2.tick_params(axis='y', labelcolor='orange')
        #
        # ax1.axhline(0, color='black', linewidth=1, linestyle='--')
        #
        # #for use_strike
        # use_strike_pts = df.dropna(subset=['use_strike'])
        # plt.scatter(use_strike_pts['time'], use_strike_pts['use_strike'], color='red', label='USE StRIKE')
        #
        # plt.title(f'{sym} Straddle ({exp})')
        # ax1.set_xlabel('Time', fontsize=5)
        #
        # time_interval = 15
        # # time_labels = [f'{i * time_interval // 60}:{(i * time_interval) % 60}' for i in range(int(375/time_interval))]
        # time_labels = df['time'].iloc[::time_interval].tolist()
        # ax1.set_xticks(range(0, len(filtered_df['time']), time_interval))
        # ax1.set_xticklabels(time_labels, fontsize=8)
        #
        # plt.setp(ax1.xaxis.get_majorticklabels(), rotation=90, ha='right')
        # ax1.grid(False)
        #
        # ax1.tick_params(axis='x', labelsize=8)
        # by_buffer = io.BytesIO()
        # plt.savefig(by_buffer, format='png')
        # plt.close()
        #
        # img = Image(by_buffer)
        # img.anchor = 'A1'
        # ws.add_image(img)
    # ----------------------------------------------------------------
        #using matplotlib
        # Convert 'time' column to datetime for filtering
        # df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S')
        # df['time'] = df['time'].dt.time
        # # Filter the dataframe based on the given time range
        # start_time = pd.to_datetime('09:19:59', format='%H:%M:%S').time()
        # end_time = pd.to_datetime('15:19:59', format='%H:%M:%S').time()
        # df_filtered = df[(df['time'] >= start_time) & (df['time'] <= end_time)]
        stt = '09:19:59'
        ett = '15:19:59'
        df_filtered = df[(df['time'] >= stt) & (df['time'] <= ett)]
        # df_filtered['time'] = df_filtered['time'].dt.time


        # Plotting
        fig, ax1 = plt.subplots(figsize=(15, 8))  # Increase the size of the graph

        # Plot forward vs time
        ax1.plot(df_filtered['time'], df_filtered['forward'], color='#FFA07A', label='Forward')  # light orange

        # Create a second y-axis for final_pnl
        ax2 = ax1.twinx()
        ax2.plot(df_filtered['time'], df_filtered['final_pnl'], color='#ADD8E6', label='Final PnL')  # light blue

        # Plot use_strike values as points and annotate
        use_strike_points = df_filtered.dropna(subset=['use_strike'])
        ax1.scatter(use_strike_points['time'], use_strike_points['use_strike'], color='#D3D3D3', label='Use Strike',
                    zorder=5)  # light gray

        # Annotate each use_strike point with its value
        for i, row in use_strike_points.iterrows():
            ax1.annotate(f"{row['use_strike']}", (row['time'], row['use_strike']),
                         textcoords="offset points", xytext=(0, 10), ha='center', fontsize=10)

        # Formatting
        ax1.set_xlabel('Time', color='black', fontsize=10)
        ax1.set_ylabel('Forward', color='black', fontsize=10)
        ax2.set_ylabel('Final PnL', color='black', fontsize=10)
        ax1.tick_params(axis='both', which='major', labelsize=8, labelcolor='black')  # Adjust tick label size for ax1
        ax2.tick_params(axis='both', which='major', labelsize=8, labelcolor='black')  # Adjust tick label size for ax2

        # settign x-axis at origin of final_pnl
        ax2.spines['bottom'].set_position(('data', 0))
        ax2.spines['bottom'].set_color('black')
        ax2.spines['bottom'].set_linewidth(1)

        plt.title(f'{sym.upper()} Straddle ({exp.upper()} Expiry)')

        # Adjust layout to make space for the legend
        # plt.tight_layout()

        by_buffer = io.BytesIO()
        plt.savefig(by_buffer, format='png')
        plt.close()

        img = Image(by_buffer)
        img.anchor = 'A1'
        ws.add_image(img)
    # ----------------------------------------------------------------
        #using plotly
        #try1
        # # Convert 'time' column to datetime for filtering
        # df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S')
        #
        # # Filter the dataframe based on the given time range
        # start_time = pd.to_datetime('09:19:59', format='%H:%M:%S').time()
        # end_time = pd.to_datetime('15:19:59', format='%H:%M:%S').time()
        # filtered_df = df[(df['time'].dt.time >= start_time) & (df['time'].dt.time <= end_time)]
        # # filtered_df = df[(df['time'] >= stt) & (df['time'] <= ett)]
        #
        # # Create the figure with Plotly
        # fig = go.Figure()
        #
        # # Add forward line plot
        # fig.add_trace(go.Scatter(x=filtered_df['time'], y=filtered_df['forward'],
        #                          mode='lines', name='Forward',
        #                          line=dict(color='orange')))
        #
        # # Add final_pnl line plot
        # fig.add_trace(go.Scatter(x=filtered_df['time'], y=filtered_df['final_pnl'],
        #                          mode='lines', name='Final PnL',
        #                          line=dict(color='lightblue'), yaxis='y2'))
        #
        # # Add use_strike points
        # use_strike_points = filtered_df.dropna(subset=['use_strike'])
        # fig.add_trace(go.Scatter(x=use_strike_points['time'], y=use_strike_points['use_strike'],
        #                          mode='markers+text', name='Use Strike',
        #                          text=use_strike_points['use_strike'],
        #                          textposition='top center',
        #                          marker=dict(color='gray', size=10)))
        #
        # # Update layout for secondary y-axis
        # fig.update_layout(
        #     title='Forward and Final PnL Over Time',
        #     xaxis=dict(title='Time', tickfont=dict(size=15)),
        #     yaxis=dict(title='Forward', tickfont=dict(size=15), titlefont=dict(color='black')),
        #     yaxis2=dict(title='Final PnL', overlaying='y', side='right', tickfont=dict(size=5),
        #                 titlefont=dict(color='black')),
        #     font=dict(size=10, color='black'),
        #     width=800,  # Adjust width
        #     height=500,  # Adjust height
        #     showlegend=False
        # )
        #
        # # Save the figure as a static image (png)
        # by_buffer = io.BytesIO()
        # fig.write_image(by_buffer, format='png')
        # by_buffer.seek(0)  # Reset buffer position to the start
        #
        # # Insert image into Excel
        # img = Image(by_buffer)
        # img.anchor = 'A1'
        # ws.add_image(img)

        # #try2
        # df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S')
        # start_time = pd.to_datetime('09:19:59', format='%H:%M:%S').time()
        # end_time = pd.to_datetime('15:19:59', format='%H:%M:%S').time()
        # filtered_df = df[(df['time'].dt.time >= start_time) & (df['time'].dt.time <= end_time)]
        #
        # fig = go.Figure()
        #
        # fig.add_trace(go.Scatter(x=filtered_df['time'], y=filtered_df['forward'],
        #                          mode='lines', name='Forward',
        #                          line=dict(color='orange')))
        # fig.add_trace(go.Scatter(x=filtered_df['time'], y=filtered_df['final_pnl'],
        #                          mode='lines', name='Final PnL',
        #                          line=dict(color='lightblue'), yaxis='y2'))
        # use_strike_points = filtered_df.dropna(subset=['use_strike'])
        # fig.add_trace(go.Scatter(x=use_strike_points['time'], y=use_strike_points['use_strike'],
        #                          mode='markers+text', name='Use Strike',
        #                          text=use_strike_points['use_strike'],
        #                          textposition='top center',
        #                          marker=dict(color='gray', size=10)))
        #
        # fig.update_layout(
        #     title='Forward and Final PnL Over Time',
        #     xaxis=dict(title='Time', tickfont=dict(size=15)),
        #     yaxis=dict(title='Forward', tickfont=dict(size=15), titlefont=dict(color='black')),
        #     yaxis2=dict(title='Final PnL', overlaying='y', side='right', tickfont=dict(size=15),
        #                 titlefont=dict(color='black')),
        #     font=dict(size=10, color='black'),
        #     width=800,
        #     height=500,
        #     showlegend=False
        # )
        #
        # by_buffer = io.BytesIO()
        # fig.write_image(by_buffer, format='png')
        # by_buffer.seek(0)
        #
        # img = Image(by_buffer)
        # img.anchor = 'A1'
        # ws.add_image(img)
    # ----------------------------------------------------------------
    # #     using seaborn
    #     # Convert 'time' column to datetime for filtering
    #     # df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S')
    #     #
    #     # # Filter the dataframe based on the given time range
    #     # start_time = pd.to_datetime('09:19:59', format='%H:%M:%S').time()
    #     # end_time = pd.to_datetime('15:19:59', format='%H:%M:%S').time()
    #     # df_filtered = df[(df['time'].dt.time >= start_time) & (df['time'].dt.time <= end_time)]
    #
    #     stt = '09:19:59'
    #     ett = '15:19:59'
    #     df_filtered = df[(df['time'] >= stt) & (df['time'] <= ett)]
    #
    #     # Plotting
    #     fig, ax1 = plt.subplots(figsize=(10, 6))  # Adjust size as needed
    #
    #     # Plot forward vs time using Seaborn
    #     sns.lineplot(x='time', y='forward', data=df_filtered, ax=ax1, color='orange', label='Forward')
    #
    #     # Create a second y-axis for final_pnl
    #     ax2 = ax1.twinx()
    #     sns.lineplot(x='time', y='final_pnl', data=df_filtered, ax=ax2, color='lightblue', label='Final PnL')
    #
    #     # Plot use_strike values as points and annotate
    #     use_strike_points = df_filtered.dropna(subset=['use_strike'])
    #     sns.scatterplot(x='time', y='use_strike', data=use_strike_points, ax=ax1, color='gray', s=100, label='Use Strike')
    #
    #     # Annotate each use_strike point with its value
    #     for i, row in use_strike_points.iterrows():
    #         ax1.annotate(f"{row['use_strike']}", (row['time'], row['use_strike']),
    #                      textcoords="offset points", xytext=(0, 10), ha='center', fontsize=10)
    #
    #     # Formatting
    #     ax1.set_xlabel('Time', color='black', fontsize=10)
    #     ax1.set_ylabel('Forward', color='black', fontsize=10)
    #     ax2.set_ylabel('Final PnL', color='black', fontsize=10)
    #     ax1.tick_params(axis='both', which='major', labelsize=5, labelcolor='black')  # Adjust tick label size for ax1
    #     ax2.tick_params(axis='both', which='major', labelsize=5, labelcolor='black')  # Adjust tick label size for ax2
    #
    #     # settign x-axis at origin of final_pnl
    #     ax2.spines['bottom'].set_position(('data', 0))
    #     ax2.spines['bottom'].set_color('black')
    #     ax2.spines['bottom'].set_linewidth(1)
    #
    #     plt.title('Forward and Final PnL Over Time')
    #
    #     # Remove legends
    #     ax1.get_legend().remove()
    #     ax2.get_legend().remove()
    #
    #     by_buffer = io.BytesIO()
    #     plt.savefig(by_buffer, format='png')
    #     plt.close()
    #
    #     img = Image(by_buffer)
    #     img.anchor = 'A1'
    #     ws.add_image(img)
# ----------------------------------------------------------------


file_status = True
while(file_status):
    filename = glob.glob(f'Data*')
    if filename:
        csv_file = os.path.join(root_dir, filename[0])
        file_status = False
    else:
        file_status = change_format()
    print(file_status)

# df = pd.read_csv(r"C:\Users\vipul\Documents\AR\for_eod\eod_data\Data_NSEFO_2024_05_28_15_32_06.csv", index_col = False)
df = pd.read_csv(csv_file, index_col=False)
# gh = df
#----------------------------------------------------------------

#take user's i/p
# sym_list = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY']
sym_list = ['BANKNIFTY']

# exp_list = ['20-Jun-24', '19-Jun-24', '25-Jun-24', '24-Jun-24']
exp_list = ['26-Jun-24']

#----------------------------------------------------------------
multiple = {
        'BANKNIFTY': {'fstrike': 100, 'round_off':100, 'delta': 15},
        'NIFTY': {'fstrike':100, 'round_off': 50, 'delta': 25},
        'FINNIFTY': {'fstrike':50, 'round_off': 50, 'delta': 40},
        'MIDCPNIFTY': {'fstrike':50, 'round_off': 25, 'delta': 75}
}
#----------------------------------------------------------------
# prev_day_data = [
#     {'NIFTY': {'use_strike': 23550, 'strd_price': 171.85, 'strd_theta':  4222142.20}},
#     {'BANKNIFTY': {'use_strike': 50300, 'strd_price': 344.15, 'strd_theta': 7921604.41}},
#     {'FINNIFTY': {'use_strike': 22600, 'strd_price': 307.17, 'strd_theta':   2100000}},
#     {'MIDCPNIFTY': {'use_strike': 12075, 'strd_price': 184.1, 'strd_theta':  2209349.72}}
# ]
prev_day_data = [
    {'BANKNIFTY': {'use_strike': 50400, 'strd_price': 841, 'strd_theta':   5100000}}
]
#----------------------------------------------------------------
tick_list = ['^NSEI', '^NSEBANK', 'NIFTY_FIN_SERVICE.NS', '^NSEMDCP50']
# tick_list = ['^NSEMDCP50']
#----------------------------------------------------------------
chk = True
sheet_list = ['intraday', 'overnight']
#implement check for sym list and exp_list

for i in range(len(sym_list)):
    break_process = False
    df_intraday = pd.DataFrame()
    df_overnight = pd.DataFrame()
    df_summary = pd.DataFrame()
    sym = sym_list[i]
    exp = exp_list[i]
    yf_tick = tick_list[i]
    num = i
    for sheet in sheet_list:
        df_fut = pd.DataFrame()
        # ---------------------------------------------------------------
        # get monthly expiry of sym = future expiry
        fut_expiry = from_master(sym)
        # ---------------------------------------------------------------
        length_count = len(df.loc[
            (df['Symbol'] == sym) & (df['Expiry'] == fut_expiry) & (df['Opttype'] == 'XX'), 'Time'])
    # ----------------------------------------------------------------------------------------------
        #to check the length of future data available so that this would not affect the creation of df of length 375
        if length_count == 375:
            df_fut['time'] = df.loc[
                (df['Symbol'] == sym) & (df['Expiry'] == fut_expiry) & (df['Opttype'] == 'XX'), 'Time']
        else:
            ffstrike_loc = df[(df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'CE') & (df['Time'] == '09:19:59')]
            ffstrike_value = int(ffstrike_loc['Strike'].mean())
            f_multiple = float(multiple[sym]['fstrike'])
            ffstrike_value_calc = int(f_multiple * round(ffstrike_value // f_multiple))
            chk_length = len(df.loc[
                (df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'CE') & (df['Strike'] == ffstrike_value_calc), 'Time'])
            if chk_length == 375:
                df_fut['time'] = df.loc[
                    (df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'CE') & (df['Strike'] == ffstrike_value_calc), 'Time']
            else:
                stt = pd.to_datetime('09:15:59')
                ett = pd.to_datetime('15:29:59')
                total_minutes = (ett - stt).total_seconds() // 60
                df_fut['time'] = pd.to_datetime(stt) + pd.to_timedelta(range(375), unit = 'm')
                df_fut['time'] = df_fut['time'].dt.strftime('%H:%M:%S')
    # ----------------------------------------------------------------------------------------------
        df_fut.reset_index()
        # ----------------------------------------------------------------
        df_fut['minute_elapsed'] = [i + 1 for i in range(375)]
        # ----------------------------------------------------------------
        val = df.loc[
            (df['Symbol'] == sym) & (df['Expiry'] == fut_expiry) & (df['Opttype'] == 'XX') & (
                        df['Time'] == '09:19:59'), 'Close'].values[0]
        df_fut['fut'] = val
        # ----------------------------------------------------------------
        if sheet == 'intraday':
            row_index = df_fut.loc[df_fut['time'] == '09:19:59'].index[0]
        else:
            row_index = df_fut.loc[df_fut['time'] == '09:15:59'].index[0]
        starting_index = row_index
        # ----------------------------------------------------------------
        f_strike = df_fut.loc[row_index, 'fut']
        f_multiple = float(multiple[sym]['fstrike'])
        fut_strike = int(f_multiple * round(f_strike // f_multiple))
    # ----------------------------------------------------------------------------------------------
        #checking if the length of the CE and PE is same or not, cause if not then the df_fut would not be made
        condition = True
        count = 0
        while condition:
            count += 1
            check_df1 = df.loc[
                (df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'CE') & (
                            df['Strike'] == fut_strike), 'Close'].values
            check_df2 = df.loc[
                (df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'PE') & (
                            df['Strike'] == fut_strike), 'Close'].values
            if len(check_df1) == 375 and len(check_df2) == 375:
                condition = False
            else:
                fut_strike += multiple[sym]['fstrike']
                condition = True

            # check to see if the src data is faulty i.e no ce pe column of any strike has 375 rows
            if count > 375:
                print(f'Cannot create table for symbol - {sym} with expiry - {exp} as src data is faulty hence skipping to next symbol')
                break_process = True
                # sys.exit()
                break
        if break_process:
            break
    # ----------------------------------------------------------------------------------------------
        df_fut['fstrike'] = fut_strike
        # ----------------------------------------------------------------
        #     BANKNIFTY29MAY2449400CE.NFO
        type_list = ['CE', 'PE']
        for each in type_list:
            ticker.append((sym + pd.to_datetime(exp).strftime('%d%b%y') + str(
                fut_strike) + each + '.NFO').upper())

        df_fut[ticker[0]] = df.loc[
            (df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'CE') & (
                        df['Strike'] == fut_strike), 'Close'].values
        df_fut[ticker[1]] = df.loc[
            (df['Symbol'] == sym) & (df['Expiry'] == exp) & (df['Opttype'] == 'PE') & (
                        df['Strike'] == fut_strike), 'Close'].values
        # ----------------------------------------------------------------
        df_fut['forward'] = df_fut['fstrike'] + df_fut[ticker[0]] - df_fut[ticker[1]]
        # ----------------------------------------------------------------
        df_fut['round_off'] = df_fut['forward'].apply(lambda x: ceiling_xcl(x, multiple[sym]['round_off']))  # this will have all the strikes
        # ----------------------------------------------------------------
        df_fut['straddle_qty'] = -50000
        # ----------------------------------------------------------------
        chk = True
        chk_row_list = []
        if sheet == 'intraday':
            range_till = 361
        else:
            range_till = 365
        for j in range(range_till):
            # ----------------------------------------------------------------
            if sheet == 'intraday':
                if chk:
                    chk_row_list.append(row_index)
                    df_fut.loc[row_index, 'use_strike'] = df_fut.loc[row_index, 'round_off']
            else:
                if chk:
                    chk_row_list.append(row_index)
                    df_fut.loc[row_index, 'use_strike'] = df_fut.loc[row_index, 'round_off']
                if row_index == starting_index:
                    # df_fut.loc[row_index, 'use_strike'] = next(data[sym]['use_strike'] for data in prev_day_data if sym in data)
                    df_fut.loc[row_index, 'use_strike'] = prev_day_data[num][sym]['use_strike']
            # ----------------------------------------------------------------
            chk_index = chk_row_list[-1]
            # ----------------------------------------------------------------
            df_fut.loc[row_index, 'ce'] = (sym + pd.to_datetime(exp).strftime('%d%b%y') + str(round(
                df_fut.loc[chk_index, 'use_strike'])) + 'CE.NFO').upper()
            df_fut.loc[row_index, 'ce_price'] = get_ce_pe_price(row_index, opt_type='CE')
            # ----------------------------------------------------------------
            df_fut.loc[row_index, 'pe'] = (sym + pd.to_datetime(exp).strftime('%d%b%y') + str(
                round(df_fut.loc[chk_index, 'use_strike'])) + 'PE.NFO').upper()
            df_fut.loc[row_index, 'pe_price'] = get_ce_pe_price(row_index, opt_type='PE')
            # ----------------------------------------------------------------
            df_fut.loc[row_index, 'straddle_price'] = df_fut.loc[row_index, 'ce_price'] + df_fut.loc[row_index, 'pe_price']
            # ----------------------------------------------------------------
              # fixing 05-jun for testing
            exp_date = datetime.datetime.strptime(exp, '%d-%b-%y').date()
            business_days_left = calc_business_days(today_date, exp_date)
            tte = business_days_left + (1 - (df_fut.loc[row_index, 'minute_elapsed']) / 375)
            df_fut.loc[row_index, 'TTE'] = (business_days_left + (1 - (df_fut.loc[row_index, 'minute_elapsed']) / 375))
            # ----------------------------------------------------------------
            # calculating option greeks
            start_time = datetime.datetime.now()
            list_for_greeks = [df_fut.loc[row_index, 'forward'], df_fut.loc[chk_row_list[-1], 'use_strike'],
                               df_fut.loc[row_index, 'TTE'], df_fut.loc[row_index, 'ce_price'],
                               df_fut.loc[row_index, 'pe_price']]
            df_fut.loc[row_index, 'ce_iv'], df_fut.loc[row_index, 'ce_delta'], df_fut.loc[row_index, 'ce_gamma'], \
            df_fut.loc[row_index, 'ce_theta'], df_fut.loc[row_index, 'ce_vega'] = get_greeks(list_for_greeks, opt_type='CE')
            # ----------------------------------------------------------------
            df_fut.loc[row_index, 'pe_iv'], df_fut.loc[row_index, 'pe_delta'], df_fut.loc[row_index, 'pe_gamma'], \
            df_fut.loc[row_index, 'pe_theta'], df_fut.loc[row_index, 'pe_vega'] = get_greeks(list_for_greeks, opt_type='PE')
            #------------------------------------------------------------------------------------------
            straddle_delta_calc = df_fut.loc[row_index, 'straddle_qty'] * (df_fut.loc[row_index, 'ce_delta'] + df_fut.loc[row_index, 'pe_delta'])
            if row_index == starting_index:
                df_fut.loc[row_index, 'strd_delta'] = straddle_delta_calc
            else:
                df_fut.loc[row_index,'strd_delta'] = straddle_delta_calc + df_fut.loc[row_index - 1,'net_fut']
            # sum_delta_trade = 0
            # if chk == True:
            #     if len(chk_row_list)>1:
            #         if row_index == (chk_row_list[-2] + 1):
            #             for index in chk_row_list[:-2]:
            #                 sum_delta_trade += df_fut.loc[index, 'delta_trade']
            #             df_fut.loc[row_index, 'strd_delta'] = straddle_delta_calc + sum_delta_trade
            #     for index in chk_row_list[:-1]:
            #         sum_delta_trade += df_fut.loc[index, 'delta_trade']
            #     df_fut.loc[row_index, 'strd_delta'] = straddle_delta_calc + sum_delta_trade
            # else:
            #     for index in chk_row_list:
            #         sum_delta_trade += df_fut.loc[index, 'delta_trade']
            #     df_fut.loc[row_index, 'strd_delta'] = straddle_delta_calc + sum_delta_trade

            # ------------------------------------------------------------------------------------------
            df_fut.loc[row_index, 'strd_gamma'] = ((df_fut.loc[row_index, 'straddle_qty'] * df_fut.loc[row_index, 'ce_gamma']) + (
                        df_fut.loc[row_index, 'straddle_qty'] * df_fut.loc[row_index, 'pe_gamma']))
            df_fut.loc[row_index, 'strd_theta'] = ((df_fut.loc[row_index, 'straddle_qty'] * df_fut.loc[row_index, 'ce_theta']) + (
                        df_fut.loc[row_index, 'straddle_qty'] * df_fut.loc[row_index, 'pe_theta']))
            df_fut.loc[row_index, 'strd_vega'] = ((df_fut.loc[row_index, 'straddle_qty'] * df_fut.loc[row_index, 'ce_vega']) + (
                        df_fut.loc[row_index, 'straddle_qty'] * df_fut.loc[row_index, 'pe_vega']))
            # ----------------------------------------------------------------
            #         if chk:
            #             df_fut.loc[row_index, 'hedge_pt'] = np.sqrt(df_fut.loc[row_index, 'theta'] / (2 * abs(df_fut.loc[row_index, 'gamma'])))

            df_fut.loc[row_index, 'hedge_pt'] = (np.sqrt(
                df_fut.loc[chk_row_list[0], 'strd_theta'] / (2 * abs(df_fut.loc[chk_row_list[0], 'strd_gamma']))))
            # ----------------------------------------------------------------
            df_fut.loc[row_index, 'D/G'] = (df_fut.loc[row_index, 'strd_delta'] / df_fut.loc[row_index, 'strd_gamma'])
            # ----------------------------------------------------------------
            condition = abs(df_fut.loc[row_index, 'D/G']) > abs(df_fut.loc[row_index, 'hedge_pt'])
            if condition:
                df_fut.loc[row_index, 'check'] = 'CHK'
            else:
                df_fut.loc[row_index, 'check'] = ''
            if sheet == 'overnight':
                if row_index == starting_index:
                    df_fut.loc[row_index, 'check'] = ''
            # ----------------------------------------------------------------
            # if (row_index == starting_index) or (str(df_fut.loc[row_index-1, 'check']) == 'CHK'):
            #     df_fut.loc[row_index, 'option_pnl'] = 0
            # else:
            #     df_fut.loc[row_index, 'option_pnl'] = df_fut.loc[row_index, 'straddle_qty'] * (
            #             df_fut.loc[row_index, 'straddle_price'] - df_fut.loc[row_index - 1, 'straddle_price'])
            if sheet == 'intraday':
                if (row_index == starting_index) or chk:
                        df_fut.loc[row_index, 'option_pnl'] = 0
                else:
                    df_fut.loc[row_index, 'option_pnl'] = round((df_fut.loc[row_index, 'straddle_qty'] * (
                                df_fut.loc[row_index,'straddle_price'] - df_fut.loc[row_index - 1,'straddle_price'])), 2)
            else:
                if (row_index == starting_index):
                    # pstrd = next(data[sym]['strd_price'] for data in prev_day_data if sym in data)
                    # df_fut.loc[row_index, 'option_pnl'] = round((df_fut.loc[row_index, 'straddle_qty'] * (
                    #         df_fut.loc[row_index, 'straddle_price'] - pstrd)), 2)
                    df_fut.loc[row_index, 'option_pnl'] = round((df_fut.loc[row_index, 'straddle_qty'] * (
                            df_fut.loc[row_index, 'straddle_price'] - prev_day_data[num][sym]['strd_price'])), 2)
                elif chk:
                    df_fut.loc[row_index, 'option_pnl'] = 0
                else:
                    df_fut.loc[row_index, 'option_pnl'] = round((df_fut.loc[row_index, 'straddle_qty'] * (
                            df_fut.loc[row_index, 'straddle_price'] - df_fut.loc[row_index - 1, 'straddle_price'])), 2)
            # ----------------------------------------------------------------
            if chk or row_index == starting_index:
                deltatrade_calc = 0
                deltatrade_calc = multiple[sym]['delta'] * round(
                    df_fut.loc[row_index, 'strd_delta'] / multiple[sym]['delta'])
                we = df_fut.loc[row_index, 'strd_delta']
                if df_fut.loc[row_index, 'strd_delta'] > 0:
                    use_multiple = -1
                else:
                    use_multiple = 1
                df_fut.loc[row_index, 'delta_trade'] = use_multiple * abs(deltatrade_calc)
            # ----------------------------------------------------------------
            sum_delta_trade = 0
            for ri in chk_row_list:
                sum_delta_trade += df_fut.loc[ri, 'delta_trade']
            df_fut.loc[row_index, 'net_fut'] = sum_delta_trade
            # ----------------------------------------------------------------
            # a, b = 0, 0
            # for i in chk_row_list:
            #     a += (df_fut.loc[i, 'delta_trade'] * df_fut.loc[i, 'forward'])
            #     b += df_fut.loc[i, 'delta_trade']

            if (row_index == starting_index):
                df_fut.loc[row_index, 'fut_pnl'] = 0
            else:
                # df_fut.loc[row_index, 'fut_pnl'] = (df_fut.loc[row_index, 'forward'] - (a / b)) * b
                # df_fut.loc[row_index, 'cum_pnl'] = df_fut['option_pnl'].sum() + df_fut.loc[row_index, 'fut_pnl']
                sp_result = sumproduct(df_fut, chk_row_list, col1='delta_trade', col2='forward')
                df_fut.loc[row_index, 'fut_pnl'] = ((df_fut.loc[row_index, 'forward'] - (sp_result/sum_delta_trade)) * df_fut.loc[row_index, 'net_fut'])
            # ----------------------------------------------------------------
            x = starting_index
            y = row_index
            cum_sum = 0
            if sheet == 'intraday':
                while x <= y:
                    cum_sum += df_fut.loc[x, 'option_pnl']
                    x += 1
                df_fut.loc[row_index, 'cum_pnl'] = (cum_sum + df_fut.loc[row_index, 'fut_pnl'])
            else:
                if row_index == starting_index:
                    df_fut.loc[row_index, 'cum_pnl'] = 0
                else:
                    while x <= y:
                        cum_sum += df_fut.loc[x, 'option_pnl']
                        x += 1
                    df_fut.loc[row_index, 'cum_pnl'] = (cum_sum + df_fut.loc[row_index, 'fut_pnl'])
            # ----------------------------------------------------------------
            if sheet == 'intraday':
                df_fut.loc[row_index, 'final_pnl'] = (df_fut.loc[row_index, 'cum_pnl'] / abs(df_fut.loc[row_index, 'straddle_qty']))
            else:
                if row_index == starting_index:
                    df_fut.loc[row_index, 'final_pnl'] = 0
                else:
                    df_fut.loc[row_index, 'final_pnl'] = (
                                df_fut.loc[row_index, 'cum_pnl'] / abs(df_fut.loc[row_index, 'straddle_qty']))
            # ----------------------------------------------------------------
            if str(df_fut.loc[row_index, 'check']) == 'CHK':
                chk = True
            else:
                chk = False
            row_index += 1

        # df_fut.loc[chk_row_list[0], 'option_pnl'], df_fut.loc[chk_row_list[0], 'fut_pnl'], df_fut.loc[
        #     chk_row_list[0], 'cum_pnl'] = 0, 0, 0
        # for each in chk_row_list:
        #     df_fut.loc[each, 'option_pnl'] = 0
        #
        # def save_to_sheet(df, sheet_name, wb):
        #     ws = wb.create_sheet(title = sheet_name)
        #     for each_line in dataframe_to_rows(df, index = False, header = True):
        #         ws.append(each_line)
        #
        #     stt = '09:19:59'
        #     ett = '15:19:59'
        #     filtered_df = df_fut[(df_fut['time'] >= stt) & (df_fut['time'] <= ett)]
        #
        #     sns.set(style='whitegrid')
        #     sns.set_context('talk')
        #
        #     # fig, ax1 = plt.subplot(figsize=(12,6))
        #     fig, ax1 = plt.subplots(figsize=(15, 7.5))
        #
        #     sns.lineplot(x='time', y='final_pnl', data=filtered_df, ax=ax1, color='blue', label='Final PnL')
        #     ax1.set_ylabel('Final PnL', color='blue', fontsize=5)
        #     ax1.tick_params(axis='y', labelcolor='blue')
        #
        #     ax2 = ax1.twinx()
        #
        #     sns.lineplot(x='time', y='forward', data=filtered_df, ax=ax2, color='orange', label='Forward')
        #     ax2.set_ylabel('Forward', color='orange', fontsize=5)
        #     ax2.tick_params(axis='y', labelcolor='orange')
        #
        #     ax1.axhline(0, color='black', linewidth=1, linestyle='--')
        #
        #     plt.title(f'{sym} Straddle ({exp})')
        #     ax1.set_xlabel('Time', fontsize=5)
        #
        #     time_interval = 15
        #     # time_labels = [f'{i * time_interval // 60}:{(i * time_interval) % 60}' for i in range(int(375/time_interval))]
        #     time_labels = df_fut['time'].iloc[::time_interval].tolist()
        #     ax1.set_xticks(range(0, len(filtered_df['time']), time_interval))
        #     ax1.set_xticklabels(time_labels, fontsize=8)
        #
        #     plt.setp(ax1.xaxis.get_majorticklabels(), rotation=90, ha='right')
        #     ax1.grid(False)
        #
        #     ax1.tick_params(axis='x', labelsize=8)
        #     by_buffer = io.BytesIO()
        #     plt.savefig(by_buffer, format='png')
        #     plt.close()
        #
        #     img = Image(by_buffer)
        #     img.anchor = 'A1'
        #     ws.add_image(img)

        df_fut = df_fut.round(decimals=3)
        # df_fut_copied = df_fut.copy()
        if sheet == 'intraday':
            df_intraday= df_fut.copy()
        else:
            df_overnight = df_fut.copy()
        # print(f'rounded df_fut is \n {df_fut}')
        # df_fut.to_csv(f'test_intraday_{sym}_{today_date}.csv')
        # print(f'CSV file name for {sym} - {today_date}')

        #plotting the static chart
        # sns.set_style('whitegrid')

        # unique_forward = df_fut['forward'].unique()
        # new_df = df_fut.reindex(df_fut.index.union(unique_forward)).sort_index()

        # stt = '09:19:59'
        # ett = '15:19:59'
        # filtered_df = df_fut[(df_fut['time'] >= stt) & (df_fut['time'] <= ett)]
        #
        # # plt.rcParams["axes.labelsize"] = 8
        # sns.set(style='whitegrid')
        # sns.set_context('talk')
        #
        # # fig, ax1 = plt.subplot(figsize=(12,6))
        # fig, ax1 = plt.subplots(figsize=(15, 7.5))
        #
        # # plotting forward on another y-axis
        # # sns.lineplot(x='time', y='forward', data=df_fut, ax=ax1, color='orange')
        # # ax1.set_ylabel('forward', color = 'orange')
        # # ax1.tick_params(axis='y', colors='orange')
        # # sns.lineplot(x='time', y='final_pnl', data=df_fut, ax=ax1, color='blue', label='Final PnL')
        # sns.lineplot(x='time', y='final_pnl', data=filtered_df, ax=ax1, color='blue', label='Final PnL')
        # ax1.set_ylabel('Final PnL', color='blue', fontsize=5)
        # ax1.tick_params(axis='y', labelcolor='blue')
        #
        # # replicating the y-axis
        # ax2 = ax1.twinx()
        #
        # #plotting final_pnl on another y-axis
        # # sns.lineplot(x='time', y='forward', data=df_fut, ax=ax2, color='orange', label='Forward')
        # sns.lineplot(x='time', y='forward', data=filtered_df, ax=ax2,color='orange', label='Forward')
        # ax2.set_ylabel('Forward', color='orange', fontsize=5)
        # ax2.tick_params(axis='y', labelcolor='orange')
        # # ax2.set_yticklabels(fontsize=8)
        #
        # # ensuring the x-axis interect y-axis at 0
        # ax1.axhline(0, color='black', linewidth = 1, linestyle = '--')
        #
        # plt.title('Final pnl and forward over time')
        # ax1.set_xlabel('Time', fontsize=5)
        # # plt.show()
        #
        # # showing time after every 5 minutes
        # # locator = mdates.MinuteLocator(interval=60)
        # # ax1.xaxis.set_major_locator(locator)
        # # ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        #
        # time_interval = 15
        # # time_labels = [f'{i * time_interval // 60}:{(i * time_interval) % 60}' for i in range(int(375/time_interval))]
        # time_labels = df_fut['time'].iloc[::time_interval].tolist()
        # ax1.set_xticks(range(0, len(filtered_df['time']), time_interval))
        # ax1.set_xticklabels(time_labels, fontsize=8)
        #
        # plt.setp(ax1.xaxis.get_majorticklabels(), rotation = 90, ha = 'right')
        # ax1.grid(False)
        #
        # ax1.tick_params(axis='x', labelsize=8)
        #
        # plt.savefig('chart.png', format='png', dpi = 1300)
        # if os.path.join(root_dir, 'chart.png'):
        #     print(f'Chart made for {sym} - {today_date}')
    # df_summary['Today'] = today_date
    # df_summary['Expiry'] = exp

    int_starting_index = df_intraday.loc[df_intraday['time'] == '09:19:59'].index[0]
    int_ending_index = df_intraday.loc[df_intraday['time'] == '15:19:59'].index[0]

    ov_starting_index = df_overnight.loc[df_overnight['time'] == '09:19:59'].index[0]
    ov_ending_index = df_overnight.loc[df_overnight['time'] == '15:19:59'].index[0]

    bod = round(df_intraday.loc[int_starting_index, 'TTE'])
    # df_summary['BoD'] = bod
    # df_summary['EoD'] = bod-1
    # df1 = yf.download(yf_tick, start=yesterday, end=today_date)
    # df2 = yf.download(yf_tick, start=today_date, end=tomorrow)
    # df_summary['Spot_Close_Yesterday'] = round(df1['Close'].values[0], 2)
    # df_summary['Spot_Close_Today'] = round(df2['Close'].values[0], 2)

    overnight_pnl = df_overnight.loc[ov_ending_index, 'final_pnl']
    overnight_pts = overnight_pnl
    # df_summary['Overnight_Pts'] = overnight_pts
    df_summary['Overnight_theta_retn'] = overnight_pts / (prev_day_data[num][sym]['strd_theta'] / 50000)
    # ptheta = next(data[sym]['strd_theta'] for data in prev_day_data if sym in data)

    intraday_pts = df_intraday.loc[int_ending_index, 'final_pnl']
    # df_summary['Intraday_Pts'] = intraday_pts
    # df_summary['Intraday_theta_retn'] = intraday_pts / (df_intraday.loc[starting_index, 'strd_theta'] / 50000)
    # df_summary['Strd_Initiation_price'] = df_intraday.loc[starting_index, 'straddle_price']
    # df_summary['Strd_Exit_price'] = df_intraday.loc[ending_index, 'straddle_price']
    eod_bod = df_overnight.loc[ov_starting_index, 'final_pnl']

    summary_data = {
        'Today': [today_date],
        'Expiry': [exp],
        'BoD': [bod],
        'EoD': [bod - 1],
        # 'Spot_Close_Yesterday': [round(df1['Close'].values[0], 2)],
        # 'Spot_Close_Today': [round(df2['Close'].values[0], 2)],
        'Overnight_Pts': [overnight_pts],
        # 'Overnight_theta_retn': [overnight_pnl / (ptheta / 50000)],
        'Overnight_theta_retn': [overnight_pnl / (prev_day_data[num][sym]['strd_theta'] / 50000)],
        'Intraday_Pts': [intraday_pts],
        'Intraday_theta_retn': [intraday_pts / (df_intraday.loc[int_starting_index, 'strd_theta'] / 50000)],
        'Strd_Initiation_price': [df_intraday.loc[int_starting_index, 'straddle_price']],
        'Strd_Exit_price': [df_intraday.loc[int_ending_index, 'straddle_price']],
        'EoD-BoD PnL': [eod_bod],
        'EoD-BoD+Intraday': [intraday_pts + eod_bod]
    }
    df_summary = df_summary.from_dict(summary_data, orient='columns')

    wb = Workbook()
    wb.remove(wb.active)
    save_to_sheet(sym, exp, df_intraday, 'Intraday', wb, plot_graph = True)
    save_to_sheet(sym, exp, df_overnight, 'Overnight', wb)
    save_to_sheet(sym, exp, df_summary, 'Summary', wb)
    filename = f'test_intraday_overnight_with_chart_{sym}_{today_date}_exp_{exp}.xlsx'
    wb.save(filename)
    print(f'Excel file created for {sym} - {today_date}')
    df_fut = df_fut.empty